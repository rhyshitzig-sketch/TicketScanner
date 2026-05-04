import io
import re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONTHS_ES = {
    1: 'ENERO',   2: 'FEBRERO',    3: 'MARZO',      4: 'ABRIL',
    5: 'MAYO',    6: 'JUNIO',      7: 'JULIO',       8: 'AGOSTO',
    9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE',
}
_MONTHS_ES_REV = {v: k for k, v in MONTHS_ES.items()}

DAYS_ES = {
    0: 'LUNES', 1: 'MARTES', 2: 'MIERCOLES', 3: 'JUEVES',
    4: 'VIERNES', 5: 'SABADO', 6: 'DOMINGO',
}
_DAYS_ES_SET = set(DAYS_ES.values())

COL_WIDTHS = {
    'C': 13.2, 'D': 20.8, 'E': 24.3, 'F': 13.0,
    'G': 13.5, 'H': 13.0, 'I': 13.0, 'J': 13.0,
    'K': 13.0, 'L': 16.0, 'M': 16.5, 'N': 12.7,
    'O': 14.3, 'P': 13.0, 'Q': 13.0, 'R': 13.0,
}

MEDIUM = Side(style='medium')
THIN   = Side(style='thin')
NO     = Side(style=None)

HEADER_FILL    = PatternFill('solid', fgColor='E7E6E6')
SEPARATOR_FILL = PatternFill('solid', fgColor='0070C0')


def write_excel(tickets, event_location='', template_bytes=None):
    if template_bytes:
        return _insert_into_template(template_bytes, tickets, event_location)

    wb = Workbook()
    wb.remove(wb.active)

    ida, regresos = [], []
    for t in tickets:
        (regresos if _is_regreso(t, event_location) else ida).append(t)

    _build_sheet(wb, 'IDA',      ida,      outbound=True)
    _build_sheet(wb, 'Regresos', regresos, outbound=False)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Template insertion ─────────────────────────────────────────────────────────

def _insert_into_template(template_bytes, tickets, event_location):
    wb = load_workbook(io.BytesIO(template_bytes))

    ida, regresos = [], []
    for t in tickets:
        (regresos if _is_regreso(t, event_location) else ida).append(t)

    for sheet_name, group, outbound in [('IDA', ida, True), ('Regresos', regresos, False)]:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            last_col = 'R' if outbound else 'Q'
            _write_title(ws, last_col)
            _write_headers(ws, outbound, last_col)
        _insert_tickets_into_sheet(wb[sheet_name], group, outbound)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _insert_tickets_into_sheet(ws, tickets, outbound):
    if not tickets:
        return
    last_col = 'R' if outbound else 'Q'

    # Sort ascending so each successive insert lands after the previous one
    for ticket in sorted(tickets, key=lambda t: (
        _parse_date_for_sort(t.get('date_departure', '')),
        _parse_time_for_sort(t.get('time_arrival', '')),
    )):
        date_str = ticket.get('date_departure', '') or 'UNKNOWN'
        sep_row  = _find_separator_row(ws, date_str)

        if sep_row is None:
            insert_at = _find_new_group_position(ws, date_str)
            if insert_at <= ws.max_row:
                _insert_rows_safe(ws, insert_at, 2)
            _write_separator(ws, insert_at, date_str, last_col)
            ws.row_dimensions[insert_at].height = 18
            _write_data_row(ws, insert_at + 1, ticket, True, True, outbound)
            ws.row_dimensions[insert_at + 1].height = 18
        else:
            end_row   = _find_group_end(ws, sep_row)
            insert_at = _find_time_position(ws, sep_row + 1, end_row, ticket.get('time_arrival', ''))
            _insert_rows_safe(ws, insert_at)
            ws.row_dimensions[insert_at].height = 18
            new_end  = _find_group_end(ws, sep_row)
            is_first = insert_at == sep_row + 1
            is_last  = insert_at == new_end
            _write_data_row(ws, insert_at, ticket, is_first, is_last, outbound)
            # Fix only the neighbouring row that changed role
            if is_first and new_end >= insert_at + 1:
                _set_row_data_borders(ws, insert_at + 1, False, insert_at + 1 == new_end, outbound)
            if is_last and insert_at - 1 >= sep_row + 1:
                _set_row_data_borders(ws, insert_at - 1, insert_at - 1 == sep_row + 1, False, outbound)


# ── Template structure detection ───────────────────────────────────────────────

def _is_separator_row(ws, row_num):
    cell = ws.cell(row=row_num, column=_col_num('C'))
    try:
        if '0070C0' in cell.fill.fgColor.rgb.upper():
            return True
    except (AttributeError, TypeError):
        pass
    val = cell.value
    if val and isinstance(val, str):
        first = val.strip().split()[0].upper()
        return first in _DAYS_ES_SET
    return False


def _find_separator_row(ws, date_str):
    target = _date_section_header(date_str).upper()
    for r in range(1, ws.max_row + 1):
        if _is_separator_row(ws, r):
            val = ws.cell(r, _col_num('C')).value or ''
            if str(val).strip().upper() == target:
                return r
    return None


def _find_group_end(ws, sep_row):
    for r in range(sep_row + 1, ws.max_row + 1):
        if _is_separator_row(ws, r):
            return r - 1
        if not any(ws.cell(r, c).value for c in range(_col_num('C'), _col_num('R') + 1)):
            return r - 1
    return ws.max_row


def _find_time_position(ws, from_row, to_row, new_time_str):
    new_time = _parse_time_for_sort(new_time_str)
    for r in range(from_row, to_row + 1):
        existing = _parse_formatted_time(str(ws.cell(r, _col_num('O')).value or ''))
        if new_time <= existing:
            return r
    return to_row + 1


def _find_new_group_position(ws, date_str):
    new_date = _parse_date_for_sort(date_str)
    year     = new_date.year if new_date != datetime.max else datetime.now().year
    for r in range(1, ws.max_row + 1):
        if _is_separator_row(ws, r):
            val = str(ws.cell(r, _col_num('C')).value or '')
            if _parse_header_to_date(val, year) > new_date:
                return r
    return ws.max_row + 1


def _set_row_data_borders(ws, row, is_first, is_last, outbound):
    top = MEDIUM if is_first else NO
    bot = MEDIUM if is_last  else NO
    for col_letter in _col_range('C', 'P'):
        ws.cell(row, _col_num(col_letter)).border = Border(
            top=top, bottom=bot,
            left=MEDIUM  if col_letter == 'C' else NO,
            right=MEDIUM if col_letter == 'P' else NO,
        )
    ws.cell(row, _col_num('Q')).border = Border(left=MEDIUM, right=NO, bottom=bot)
    if outbound:
        ws.cell(row, _col_num('R')).border = Border(left=NO, right=MEDIUM, bottom=bot)


def _insert_rows_safe(ws, row_idx, amount=1):
    """Insert rows and correctly shift all merged cell ranges."""
    affected = [
        (r.min_row, r.min_col, r.max_row, r.max_col)
        for r in list(ws.merged_cells.ranges)
        if r.max_row >= row_idx
    ]
    for min_row, min_col, max_row, max_col in affected:
        ws.unmerge_cells(
            f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        )
    ws.insert_rows(row_idx, amount)
    for min_row, min_col, max_row, max_col in affected:
        new_min = min_row + amount if min_row >= row_idx else min_row
        new_max = max_row + amount
        ws.merge_cells(start_row=new_min, start_column=min_col,
                       end_row=new_max, end_column=max_col)


# ── Date/time parsing helpers ──────────────────────────────────────────────────

def _parse_date_for_sort(date_str):
    try:
        return datetime.strptime(date_str, '%d/%m/%Y')
    except ValueError:
        return datetime.max


def _parse_time_for_sort(time_str):
    m = re.match(r'^(\d{1,2}):(\d{2})$', (time_str or '').strip())
    return (int(m.group(1)), int(m.group(2))) if m else (99, 99)


def _parse_formatted_time(time_str):
    """Parse the 'HHhMM' format used in the output cells back to a sortable tuple."""
    m = re.match(r'^(\d{1,2})[Hh](\d{2})$', time_str.strip())
    return (int(m.group(1)), int(m.group(2))) if m else (99, 99)


def _parse_header_to_date(header, year=None):
    """Parse a separator label like 'LUNES 21 DE MAYO' back to a datetime."""
    m = re.match(r'\w+\s+(\d+)\s+DE\s+(\w+)', header.strip().upper())
    if m:
        day   = int(m.group(1))
        month = _MONTHS_ES_REV.get(m.group(2))
        if month:
            try:
                return datetime(year or datetime.now().year, month, day)
            except ValueError:
                pass
    return datetime.max


# ── Sheet builder (from scratch) ───────────────────────────────────────────────

def _build_sheet(wb, name, tickets, outbound):
    ws       = wb.create_sheet(name)
    last_col = 'R' if outbound else 'Q'

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    _write_title(ws, last_col)
    _write_headers(ws, outbound, last_col)

    cur_row = 4
    groups  = _group_by_date(tickets)
    for date_str, group in sorted(groups.items(), key=lambda x: _sort_key(x[0])):
        _write_separator(ws, cur_row, date_str, last_col)
        cur_row += 1
        group_sorted = sorted(group, key=lambda t: _parse_time_for_sort(t.get('time_arrival', '')))
        for idx, ticket in enumerate(group_sorted):
            _write_data_row(ws, cur_row, ticket, idx == 0, idx == len(group_sorted) - 1, outbound)
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1


# ── Cell writers ───────────────────────────────────────────────────────────────

def _write_title(ws, last_col):
    ws.merge_cells(f'C1:{last_col}2')
    cell           = ws['C1']
    cell.value     = 'WOW CONTROL VIAJES'
    cell.font      = Font(bold=True, size=22, color='000000')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28


def _write_headers(ws, outbound, last_col):
    labels = {
        'C': 'PUESTO',    'D': 'NOMBRE',     'E': 'APELLIDOS',
        'F': 'PASAPORTE', 'G': 'AEROLÍNEA',  'H': 'Nº VUELO',
        'I': 'LOCALIZADOR',
        'J': 'VUELO IDA' if outbound else 'VUELO REGRESO',
        'K': 'ESCALA',    'L': 'FECHA SALIDA', 'M': 'FECHA LLEGADA',
        'N': 'HORA SALIDA', 'O': 'HORA LLEGADA', 'P': 'TERMINAL',
        'Q': 'EMITIDO' if outbound else 'CHECK IN?',
        'R': 'CHECK IN?',
    }
    for col_letter in _col_range('C', last_col):
        col_num        = _col_num(col_letter)
        cell           = ws.cell(row=3, column=col_num)
        cell.value     = labels.get(col_letter, '')
        cell.font      = Font(bold=True, color='000000')
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = Border(
            top=MEDIUM, bottom=MEDIUM,
            left=MEDIUM if col_letter == 'C' else THIN,
            right=MEDIUM if col_letter == last_col else THIN,
        )
    ws.row_dimensions[3].height = 18


def _write_separator(ws, row, date_str, last_col):
    ws.merge_cells(f'C{row}:{last_col}{row}')
    cell           = ws[f'C{row}']
    cell.value     = _date_section_header(date_str)
    cell.font      = Font(bold=True, size=12, color='FFFFFF')
    cell.fill      = SEPARATOR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = Border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM)
    ws.row_dimensions[row].height = 18


def _write_data_row(ws, row, ticket, is_first, is_last, outbound):
    vals = _format_ticket(ticket)
    data = {
        'D': vals['first_name'],   'E': vals['last_name'],
        'G': vals['airline'],      'H': vals['flight_number'],
        'I': vals['confirmation'], 'J': vals['route'],
        'L': vals['date_departure'], 'M': vals['date_arrival'],
        'N': vals['time_departure'], 'O': vals['time_arrival'],
    }

    top = MEDIUM if is_first else NO
    bot = MEDIUM if is_last  else NO

    for col_letter in _col_range('C', 'P'):
        cell           = ws.cell(row=row, column=_col_num(col_letter))
        cell.value     = data.get(col_letter, '')
        cell.font      = Font(bold=(col_letter in ('M', 'O')), color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = Border(
            top=top, bottom=bot,
            left=MEDIUM  if col_letter == 'C' else NO,
            right=MEDIUM if col_letter == 'P' else NO,
        )

    ws.cell(row=row, column=_col_num('Q')).border = Border(
        left=MEDIUM, right=NO, bottom=bot,
    )
    if outbound:
        ws.cell(row=row, column=_col_num('R')).border = Border(
            left=NO, right=MEDIUM, bottom=bot,
        )


# ── Formatting helpers ─────────────────────────────────────────────────────────

def _format_ticket(t):
    return {
        'first_name':     t.get('first_name',    '').strip().upper(),
        'last_name':      t.get('last_name',     '').strip().upper(),
        'airline':        _fmt_airline(t.get('airline', '')),
        'flight_number':  t.get('flight_number', '').strip().upper(),
        'confirmation':   t.get('confirmation',  '').strip().upper(),
        'route':          _fmt_route(t.get('route', '')),
        'date_departure': _fmt_date(t.get('date_departure', '')),
        'date_arrival':   _fmt_date(t.get('date_arrival',   '')),
        'time_departure': _fmt_time(t.get('time_departure', '')),
        'time_arrival':   _fmt_time(t.get('time_arrival',   '')),
    }


def _fmt_airline(v):
    v = v.strip().upper()
    return {'AIR EUROPA': 'AIREUROPA'}.get(v, v)


def _fmt_route(v):
    return re.sub(r'\s*-\s*', '-', v.strip().upper())


def _fmt_date(v):
    v = v.strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/\d{4}$', v)
    if not m:
        return v.upper()
    month = MONTHS_ES.get(int(m.group(2)), '')
    return f'{int(m.group(1))} DE {month}' if month else v


def _fmt_time(v):
    v = v.strip()
    m = re.match(r'^(\d{1,2}):(\d{2})$', v)
    if not m:
        return v.upper()
    return f'{int(m.group(1)):02d}H{m.group(2)}'


def _date_section_header(date_str):
    try:
        dt = datetime.strptime(date_str, '%d/%m/%Y')
        return f'{DAYS_ES[dt.weekday()]} {dt.day} DE {MONTHS_ES[dt.month]}'
    except ValueError:
        return date_str.upper()


# ── Route direction detection ──────────────────────────────────────────────────

def _is_regreso(ticket, event_location=''):
    route = ticket.get('route', '').upper()
    parts = [p.strip() for p in re.split(r'\s*-\s*', route) if p.strip()]
    names = _event_names(event_location)
    return bool(parts) and _matches(parts[0], names)


def _event_names(event_location):
    names = {p.strip().upper() for p in re.split(r'[,/;|]', event_location) if p.strip()}
    return names or {'LIS', 'LISBOA', 'LISBON'}


def _matches(value, names):
    value = value.strip().upper()
    return any(value == n or n in value for n in names)


# ── Utilities ──────────────────────────────────────────────────────────────────

def _group_by_date(tickets):
    groups = defaultdict(list)
    for t in tickets:
        groups[t.get('date_departure', '') or 'UNKNOWN'].append(t)
    return groups


def _sort_key(date_str):
    try:
        return datetime.strptime(date_str, '%d/%m/%Y')
    except ValueError:
        return datetime.max


def _col_range(start, end):
    return [chr(c) for c in range(ord(start), ord(end) + 1)]


def _col_num(col_letter):
    return ord(col_letter) - ord('A') + 1
